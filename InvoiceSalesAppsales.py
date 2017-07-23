from tkinter import *
import openpyxl as xl
from pypodio2 import api
from pathlib import Path


class Window(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)

        self.myPodioApp = master

        self.init_window()

    def init_window(self):
        self.myPodioApp.title("QUICK EFFICIENT FAST SPEEDX1000 EL RAPIDO AMAZING SALES PROGRAM 2000")
        self.pack(fill=BOTH, expand=1)

        self.error = Label(self, text="Felmeddelande: ")
        exists = True

        def selectItemInListbox(event):
            print("called")
            entries[3].delete(0, END)
            listBox = event.widget
            listOfDirectoriesSelection = listOfDirectories.curselection()
            listOfDirectoriesSelectionGet = listBox.get(listOfDirectoriesSelection[0])
            entries[3].insert(0, listOfDirectoriesSelectionGet.rstrip())

        def defineErrorMessage(message):
            destroyErrorMessage()
            self.error = Label(self, text="Felmeddelande: " + message)
            self.error.place(x=buttonXValues[1] + 40, y=buttonYValues[1] + topMostElementsY)
            print("created")

        def destroyErrorMessage():
            self.error.destroy()
            print("destroyed")

        def showHidePassword():
            if showHideVar.get() == False:
                entries[1].config(show="*")
            else:
                entries[1].config(show="")

        def storeLoginInfo():
            global emailAddress
            global userPassword

            emailAddress = entries[0].get()
            userPassword = entries[1].get()

            print(emailAddress)
            print(userPassword)

        def findCompany():
            global compName
            global orgNumber
            global phNumber
            global location
            global name

            compName = entries[2].get()
            sPath = entries[3].get()
            sPathList = list(sPath)
            lastCharacterInPathIndex = len(sPathList) - 1  # first character has index value 0, like lists
            z = 0

            while z <= lastCharacterInPathIndex:
                if sPathList[z] == '\\':
                    sPathList[z] = '/'

                z += 1
                print(z)

            sPath = ''.join(sPathList)
            print(sPath)
            wb = xl.load_workbook(
                sPath)  # C:/Users/Nils/Documents/uppgifter.xlsx / C:\Users\Nils\Documents\uppgifter.xlsx
            sheet1 = wb.worksheets[
                0]  # förutsätter att vi vill söka i blad med indexnummer 0 (indexnumren förändras inte när man byter plats på dem i excel-dokumentet)
            sPathWithNewLine = sPath + '\n'

            saveFilePath = P ath('C:/Users/Nils/Documents/sparfilPodio.txt')
            if saveFilePath.exists():
                saveFileAppend = open('C:/Users/Nils/Documents/sparfilPodio.txt', 'a')
                saveFileRead = open('C:/Users/Nils/Documents/sparfilPodio.txt', 'r')
                saveFileLines = saveFileRead.readlines()
                saveFileLinesAsList = []
                x = 0
                alreadySaved = False

                for line in saveFileLines:
                    print(saveFileLines[x])
                    saveFileLinesAsList.append(saveFileLines[x])
                    if saveFileLinesAsList[x] == sPathWithNewLine:
                        print('dir already used, no need to save it again\n')
                        alreadySaved = True
                    x += 1

                insertPathToLastPosition = len(saveFileLinesAsList)

                print(alreadySaved)
                if alreadySaved == False:
                    lengthOfList = 0
                    for line in saveFileLines:
                        lengthOfList += 1
                        print(lengthOfList)

                    if lengthOfList <= 3:
                        saveFileAppend.write(sPathWithNewLine)
                        print('saved new path in save file')
                        listOfDirectories.insert(insertPathToLastPosition, sPath)
                    else:
                        del saveFileLinesAsList[0]

                        saveFileLinesAsList.append(sPathWithNewLine)
                        saveFileWrite = open('C:/Users/Nils/Documents/sparfilPodio.txt', 'w')
                        saveFileWrite.write(''.join(saveFileLinesAsList))
                        print('removed first path of the list and added the new path to the save-file')
                        listOfDirectories.delete(0)
                        listOfDirectories.insert(insertPathToLastPosition, sPath)
            else:
                saveFile = open('C:/Users/Nils/Documents/sparfilPodio.txt', 'w')
                saveFile.write(sPathWithNewLine)
                print('created save-file')
                listOfDirectories.insert(1, sPath)

            for x in range(5):
                columns[x].delete(1.0, END)

            x = 1
            y = sheet1.max_row
            empty = True

            print('y = ' + str(y))
            while x <= y:
                if sheet1.cell(row=x, column=2).value == compName:
                    columns[0].insert(0.0, compName)

                    orgNumber = str(sheet1.cell(row=x, column=3).value)
                    columns[1].insert(0.0, orgNumber)

                    phNumberList = list(str(sheet1.cell(row=x, column=6).value))

                    if phNumberList[0] == '0':
                        phNumber = ''.join(phNumberList)
                        columns[2].insert(0.0, phNumber)
                    else:
                        phNumberList.insert(0, '0')
                        phNumber = ''.join(phNumberList)
                        columns[2].insert(0.0, phNumber)

                    location = sheet1.cell(row=x, column=5).value
                    columns[3].insert(0.0, location)

                    name = sheet1.cell(row=x, column=4).value
                    columns[4].insert(0.0, name)

                    x = y
                    empty = False

                print('x = ' + str(x))
                x += 1

            if empty == True:
                defineErrorMessage("Hittade inte företaget i filen")
            else:
                destroyErrorMessage()
                empty = True

        def approveAndCreate():
            client_id  = 'fors'
            client_secret = '9W0pxSqKo61LG4qTToZrw9Dc41ICbtppjq1066yrt5i6bwKBx9nagxqCZi9VU19f'
            username = emailAddress
            password = userPassword

            searchFor = {"query": compName, "ref_type": 'item', "search_fields": 'title'}

            phoneNumberString = str(phNumber)
            phoneArray = [phoneNumberString]
            contactValues = {'name': name, 'phone': phoneArray}

            podio = api.OAuthClient(
                client_id,
                client_secret,
                username,
                password,
            )

            found = podio.Search.searchApp(8014787, searchFor)
            foundString = str(found)
            foundRightString = foundString.find(compName)
            contactValues = {'name': name, 'phone': phoneArray}

            if foundRightString == -1:
                destroyErrorMessage()
                profileID = podio.Contact.create(2209982, contactValues)
                values = {'fields': {62123631: compName,
                                     62124685: orgNumber,
                                     62124688: [122654776, 177185617],
                                     62124687: profileID['profile_id'],
                                     62124690: 8,
                                     62124689: 1,
                                     62124686: location,
                                     62128871: 18
                                     }}
                podio.Item.create(8014787, values)
                print(foundRightString)

            else:
                defineErrorMessage("Företaget finns redan i Podio")
                print(foundString)
                print(foundRightString)

        # Various x- & y-constants
        leftmostElementsX = 50
        topMostElementsY = 5

        labelsY = 180

        companyNameAndPathY = 70

        textboxY = 150
        textBoxWidth = 30
        textboxXDistance = textBoxWidth * 8 + 10

        emailAndPasswordX = 800
        emailAndPasswordY = topMostElementsY
        emailAndPasswordYDistance = 25
        emailAndPasswordXAxisMultiplier = 2

        instructionXValues = leftmostElementsX + textboxXDistance * 4 - 50
        instructionValuesYDistance = 30
        instructionYValuesMultiplier = 5

        # Various lists
        columns = []
        columnLabels = []
        columnStrings = ['Företagets namn', 'Organisationsnummer', 'Telefonnummer', 'Plats', 'Namn på kontakt']

        entries = []
        entryLabels = []
        entrylabelStrings = ['Mailadress', 'Lösenord', 'Företagets namn', 'I Excel-fil (inkl. sökväg)']
        entryXValues = [leftmostElementsX + textboxXDistance * emailAndPasswordXAxisMultiplier, leftmostElementsX]
        buttons = []

        buttonTexts = ['Spara', 'Hitta', 'Godkänn']
        buttonXValues = [leftmostElementsX + textboxXDistance * emailAndPasswordXAxisMultiplier, entryXValues[1], 525]
        buttonYValues = [110, 110, 220]
        buttonCommands = {0: storeLoginInfo, 1: findCompany, 2: approveAndCreate}
        showConfig = ['', '*', '', '']

        instructionLabels = []
        instructionLabelStrings = ['1. Skriv in dina inloggningsuppgifter och spara dessa',
                                   '2. Skriv in företagets namn för att hitta företaget',
                                   '3. Kontrollera att kolumnerna hamnat rätt och godkänn']
        instructionYValues = [topMostElementsY * instructionYValuesMultiplier,
                              topMostElementsY * instructionYValuesMultiplier + instructionValuesYDistance,
                              topMostElementsY * instructionYValuesMultiplier + instructionValuesYDistance * 2]

        # Creating the text boxes
        for x in range(5):
            columns.append(x)
            columns[x] = Text(width=textBoxWidth, height=1)
            columns[x].insert(0.0, "")
            columns[x].place(x=leftmostElementsX + textboxXDistance * x, y=textboxY)

            columnLabels.append(x)
            columnLabels[x] = Label(text=columnStrings[x])
            columnLabels[x].place(x=leftmostElementsX + textboxXDistance * x, y=labelsY)

        # Creating the entry boxes
        c = 0
        d = 0
        for a in range(2):
            for b in range(2):
                entries.append(d)
                entries[d] = Entry(show=showConfig[d], width=40)
                entries[d].place(x=entryXValues[a], y=emailAndPasswordY + emailAndPasswordYDistance * (c + 1))

                entryLabels.append(d)
                entryLabels[d] = Label(text=entrylabelStrings[d])
                entryLabels[d].place(x=entryXValues[a], y=emailAndPasswordY + emailAndPasswordYDistance * c)
                c += 2
                d += 1
            c = 0

        # Creating the buttons
        for x in range(3):
            buttons.append(x)
            buttons[x] = Button(self, text=buttonTexts[x], command=buttonCommands[x])
            buttons[x].place(x=buttonXValues[x], y=buttonYValues[x])
        buttons[2].config(height=1, width=40)

        # Creating the show/hide password check-button
        showHideVar = BooleanVar()
        showHidePasswordCheckButton = Checkbutton(self, text="Göm/visa lösenord", variable=showHideVar, onvalue=True,
                                                  offvalue=False, command=showHidePassword)
        showHidePasswordCheckButton.place(
            x=leftmostElementsX + textboxXDistance * (emailAndPasswordXAxisMultiplier + 1),
            y=emailAndPasswordY + emailAndPasswordYDistance * 3)

        # Creating instructing labels
        for x in range(3):
            instructionLabels.append(x)
            instructionLabels[x] = Label(text=instructionLabelStrings[x])
            instructionLabels[x].place(x=instructionXValues, y=instructionYValues[x])

        # Creating the listbox
        listOfDirectories = Listbox(self)
        listOfDirectories.config(width=40, height=4)
        listOfDirectories.place(x=leftmostElementsX + textboxXDistance,
                                y=emailAndPasswordY + emailAndPasswordYDistance * 1)
        saveFilePath = Path('C:/Users/Nils/Documents/sparfilPodio.txt')
        if saveFilePath.exists():
            listOfDirectoriesFile = open('C:/Users/Nils/Documents/sparfilPodio.txt', 'r')
            listOfDirectoriesLines = listOfDirectoriesFile.readlines()
            listOfDirectoriesLinesList = []
            y = 1
            x = 0
            for line in listOfDirectoriesLines:
                listOfDirectories.insert(y, listOfDirectoriesLines[x])
                x += 1
                y += 1
            listOfDirectories.bind("<Double-Button-1>", selectItemInListbox)
        listOfDirectoriesLabel = Label(text='Senast använda filer')
        listOfDirectoriesLabel.place(x=leftmostElementsX + textboxXDistance, y=emailAndPasswordY)

root = Tk()
root.geometry("1350x300")
app = Window(root)
root.mainloop()